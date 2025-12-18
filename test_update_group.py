import http.client
import json
import time
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

# -------------------------- 配置项 --------------------------
NEW_GROUP = "限时特价,优质gemini"  # 新的 group 值
API_DOMAIN = "api.vectorengine.ai"
API_PATH = "/api/token/"
HEADERS_BASE = {
    'new-api-user': '142538',
    'Authorization': 'SxpO4tsw05gn5icEKyBb4iSKfE/Y3TEj',
    'content-type': 'application/json'
}
EXCEL_READ_PATH = r"E:\PythonProject\sigma\20251217_令牌列表_400.xlsx"  # 读取的Excel文件路径

# ⚠️ 并发设置
MAX_WORKERS = 10  # 建议设置 5-10，设置过高容易触发 API 的 429 限流

# -------------------------- 核心功能函数 --------------------------

def get_token_info(token_key, token_name=None):
    """
    通过令牌key或名称查询令牌完整信息
    注意：此函数内新建连接，是线程安全的
    """
    conn = None
    try:
        conn = http.client.HTTPSConnection(API_DOMAIN, timeout=10)
        headers = {
            'new-api-user': HEADERS_BASE['new-api-user'],
            'Authorization': HEADERS_BASE['Authorization'],
        }
        
        # 去掉 sk- 前缀进行匹配
        key_without_prefix = token_key.replace("sk-", "") if token_key.startswith("sk-") else token_key
        
        # 1. 优先尝试：通过名称搜索（效率最高）
        if token_name:
            # URL 编码处理 (简单处理，如果名字有特殊符号最好用 urllib.parse.quote)
            safe_name = token_name.replace(" ", "%20")
            search_url = f"/api/token/search?keyword={safe_name}"
            conn.request("GET", search_url, headers=headers)
            res = conn.getresponse()
            data = json.loads(res.read().decode("utf-8"))
            
            # 必须重新建立连接或读取完数据，http.client同一连接不能连续请求
            conn.close() 

            if data.get("success") and data.get("data"):
                token_list = data["data"]
                if isinstance(token_list, list):
                    for token in token_list:
                        if token.get("name") == token_name or token.get("key") == key_without_prefix:
                            return token

        # 2. 备选方案：分页获取（效率较低，仅当搜索失败时使用）
        # 注意：在多线程下大量调用分页可能会对服务器造成压力
        page = 0
        while True:
            conn = http.client.HTTPSConnection(API_DOMAIN, timeout=10) # 重新连接
            conn.request("GET", f"/api/token/?p={page}&size=100", headers=headers)
            res = conn.getresponse()
            data = json.loads(res.read().decode("utf-8"))
            conn.close()
            
            if not data.get("success"):
                break
            
            response_data = data.get("data", {})
            token_list = response_data.get("items", []) if isinstance(response_data, dict) else (response_data if isinstance(response_data, list) else [])
            
            if not token_list:
                break

            for token in token_list:
                if (token.get("key") == key_without_prefix or 
                    token.get("key") == token_key or
                    (token_name and token.get("name") == token_name)):
                    return token
            
            if len(token_list) < 100 or page > 10:  # 限制最多翻10页
                break
            page += 1
        
        return None
    except Exception as e:
        # print(f"查询出错: {e}") # 调试用
        return None
    finally:
        if conn:
            conn.close()

def update_token_group_request(token_info, new_group):
    """发送更新请求"""
    conn = None
    try:
        # 构造 Payload
        payload = json.dumps({
            "id": token_info["id"],
            "name": token_info.get("name", ""),
            "remain_quota": token_info.get("remain_quota", 0),
            "expired_time": token_info.get("expired_time", -1),
            "unlimited_quota": token_info.get("unlimited_quota", False),
            "model_limits_enabled": token_info.get("model_limits_enabled", False),
            "model_limits": token_info.get("model_limits", ""),
            "group": new_group,  # ✅ 核心修改点
            "mj_image_mode": token_info.get("mj_image_mode", "default"),
            "mj_custom_proxy": token_info.get("mj_custom_proxy", ""),
            "selected_groups": token_info.get("selected_groups", []),
            "allow_ips": token_info.get("allow_ips", "")
        })

        conn = http.client.HTTPSConnection(API_DOMAIN, timeout=10)
        conn.request("PUT", API_PATH, payload, HEADERS_BASE)
        res = conn.getresponse()
        
        if res.status not in [200, 201]:
            return False, f"HTTP {res.status}"
            
        data = json.loads(res.read().decode("utf-8"))
        if data.get("success"):
            return True, "OK"
        else:
            return False, data.get('message', 'API返回失败')
            
    except Exception as e:
        return False, str(e)
    finally:
        if conn:
            conn.close()

# -------------------------- 线程任务函数 --------------------------

def process_single_token(token_data):
    """
    单个线程的处理逻辑
    返回格式: (状态码, token_name, 消息)
    状态码: 0=失败, 1=成功, 2=跳过
    """
    token_name = token_data["name"]
    token_key = token_data["key"]
    
    try:
        # 1. 查询信息
        token_info = get_token_info(token_key, token_name=token_name)
        
        if not token_info:
            return (0, token_name, "无法获取令牌信息（可能不存在或API超时）")
            
        # 2. 检查是否需要更新
        current_group = token_info.get("group", "")
        if current_group == NEW_GROUP:
            return (2, token_name, "Group已匹配，无需更新")
            
        # 3. 执行更新
        success, msg = update_token_group_request(token_info, NEW_GROUP)
        
        if success:
            return (1, token_name, f"更新成功: {current_group} -> {NEW_GROUP}")
        else:
            return (0, token_name, f"更新请求失败: {msg}")

    except Exception as e:
        return (0, token_name, f"处理异常: {str(e)}")

# -------------------------- Excel 读取 --------------------------
def load_tokens_from_excel(excel_path):
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        tokens = []
        for row in range(2, ws.max_row + 1):
            name = ws[f"A{row}"].value
            key = ws[f"B{row}"].value
            if name and key:
                tokens.append({"name": str(name), "key": str(key)})
        return tokens
    except Exception as e:
        print(f"❌ 读取Excel失败: {e}")
        return []

# -------------------------- 主程序 --------------------------
if __name__ == "__main__":
    print(f"正在读取 {EXCEL_READ_PATH} ...")
    tokens = load_tokens_from_excel(EXCEL_READ_PATH)
    
    if not tokens:
        print("❌ 未找到有效数据，程序结束。")
        exit()
        
    print(f"找到 {len(tokens)} 个令牌。")
    print(f"目标 Group: {NEW_GROUP}")
    print(f"线程数: {MAX_WORKERS}")
    print("=" * 60)
    
    success_count = 0
    fail_count = 0
    skip_count = 0
    
    start_time = time.time()
    
    # 建立线程池
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # 提交所有任务
        future_to_token = {executor.submit(process_single_token, t): t for t in tokens}
        
        processed_count = 0
        total_count = len(tokens)
        
        for future in as_completed(future_to_token):
            processed_count += 1
            status, name, msg = future.result()
            
            # 状态码: 0=失败, 1=成功, 2=跳过
            if status == 1:
                success_count += 1
                print(f"[{processed_count}/{total_count}] ✅ {name} | {msg}")
            elif status == 2:
                skip_count += 1
                # 跳过的信息可以打印得简洁一点，或者注释掉不打印
                print(f"[{processed_count}/{total_count}] ⏭️  {name} | {msg}")
            else:
                fail_count += 1
                print(f"[{processed_count}/{total_count}] ❌ {name} | {msg}")

    duration = time.time() - start_time
    
    print("=" * 60)
    print(f"全部处理完成！耗时: {duration:.2f}秒")
    print(f"✅ 更新成功: {success_count}")
    print(f"⏭️  跳过(无需更新): {skip_count}")
    print(f"❌ 失败: {fail_count}")