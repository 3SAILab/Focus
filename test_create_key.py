import http.client
import json
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed

# -------------------------- 配置项 --------------------------
QUOTA = 1.75  # 配额值
CREATE_COUNT = 400  # 要创建的令牌数量
API_DOMAIN = "api.vectorengine.ai"
API_PATH = "/api/token/"
HEADERS_BASE = {
    'new-api-user': '142538',
    'Authorization': 'SxpO4tsw05gn5icEKyBb4iSKfE/Y3TEj',
    'content-type': 'application/json'
}

# ⚠️ 并发线程数：建议设置在 5-20 之间。太高可能会导致 429 报错或被封号
MAX_WORKERS = 10 

# 获取当前日期字符串 (全局变量，用于文件名和令牌名)
current_date = datetime.now().strftime("%Y%m%d")

# 文件名配置
EXCEL_SAVE_PATH = f"{current_date}_令牌列表_{CREATE_COUNT}.xlsx"

# -------------------------- 初始化Excel --------------------------
wb = Workbook()
ws = wb.active
ws.title = "令牌列表"
ws['A1'] = "令牌名称(name)"
ws['B1'] = "令牌密钥(key)"

header_font = Font(bold=True, color="FFFFFF")
header_fill_style = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for cell in ws[1]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.fill = header_fill_style

# -------------------------- 核心逻辑函数 --------------------------
def get_payload(quota, n):
    """生成单个令牌的请求参数"""
    str_quota = str(quota).replace(".", "_")
    
    # ✅ 修改点：名称现在是 动态日期 + 动态总量 + 配额 + 序号(n+1)
    # 例如：20231217_400_1_75rmb_1
    name = f"{current_date}_{CREATE_COUNT}_{str_quota}rmb_{n+1}"
    
    payload = json.dumps({
        "name": name,
        "remain_quota": int(quota * 1000000),
        "expired_time": -1,
        "unlimited_quota": False,
        "model_limits_enabled": False,
        "model_limits": "",
        "group": "限时特价,default",
        "mj_image_mode": "default",
        "mj_custom_proxy": "",
        "selected_groups": [],
        "allow_ips": ""
    })
    return payload, name

def create_token_task(n):
    """
    单个线程执行的任务函数
    返回: (True, name, key) 或 (False, error_msg, None)
    """
    conn = None
    try:
        payload, token_name = get_payload(QUOTA, n)
        
        # 建立连接
        conn = http.client.HTTPSConnection(API_DOMAIN, timeout=10)
        conn.request("POST", API_PATH, payload, HEADERS_BASE)
        res = conn.getresponse()
        
        if res.status not in [200, 201]:
            return False, f"HTTP {res.status}: {res.reason}", None

        data_str = res.read().decode("utf-8")
        data = json.loads(data_str)

        if data.get("success"):
            token_key = data.get('data')
            return True, token_name, token_key
        else:
            return False, f"API Error: {data.get('message')}", None

    except Exception as e:
        return False, str(e), None
    finally:
        if conn:
            conn.close()

# -------------------------- 主程序执行 --------------------------
if __name__ == "__main__":
    print(f"🚀 开始多线程创建 {CREATE_COUNT} 个令牌...")
    print(f"📅 当前日期: {current_date}")
    print(f"⚙️  线程数: {MAX_WORKERS} | 保存路径: {EXCEL_SAVE_PATH}")
    print("-" * 50)

    success_count = 0
    fail_count = 0
    row = 2

    start_time = time.time()

    # 使用线程池管理器
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # 提交所有任务
        future_tasks = {executor.submit(create_token_task, n): n for n in range(CREATE_COUNT)}

        # as_completed 会在某个任务完成时立即 yield 结果
        for future in as_completed(future_tasks):
            success, result_1, result_2 = future.result()
            
            if success:
                token_name = result_1
                token_key = result_2
                
                # ✅ 在主线程写入 Excel
                ws[f"A{row}"] = token_name
                ws[f"B{row}"] = token_key
                row += 1
                success_count += 1
                
                # 打印简略进度
                if success_count % 10 == 0 or success_count == CREATE_COUNT:
                    print(f"✅ 进度: {success_count}/{CREATE_COUNT} (最新: {token_name})")
            else:
                fail_count += 1
                error_msg = result_1
                print(f"❌ 创建失败: {error_msg}")

    end_time = time.time()
    duration = end_time - start_time

    # -------------------------- 保存与统计 --------------------------
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50

    print("=" * 50)
    print(f"处理完毕！耗时: {duration:.2f}秒")
    print(f"✅ 成功：{success_count}")
    print(f"❌ 失败：{fail_count}")

    try:
        print(f"💾 正在保存文件...")
        wb.save(EXCEL_SAVE_PATH)
        print(f"📄 文件已保存至：{EXCEL_SAVE_PATH}")
    except Exception as e:
        print(f"❌ 保存Excel失败: {e}")
        print("请关闭文件后重试。")