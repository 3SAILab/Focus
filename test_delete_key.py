import http.client
import json
import time
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

# -------------------------- 配置项 --------------------------
# 目标 Excel 文件路径
EXCEL_FILE_PATH = "删除列表.xlsx"

# 要读取的列（'A'代表第一列，如果ID在B列请改为'B'）
TARGET_COLUMN = 'A' 

# 是否有表头（True则跳过第一行，False则从第一行开始读）
HAS_HEADER = True 

# API 配置
API_DOMAIN = "api.vectorengine.ai"
HEADERS = {
    'new-api-user': '142538',  # 请替换为您的用户ID
    'Authorization': 'SxpO4tsw05gn5icEKyBb4iSKfE/Y3TEj', # 请替换为您的Auth Key
    'content-type': 'application/json'
}

# 并发线程数
MAX_WORKERS = 10

# -------------------------- 核心逻辑 --------------------------

def delete_token_task(token_id):
    """
    单个删除任务
    """
    conn = None
    try:
        # 确保ID是纯净的字符串/数字，去除可能的空格
        clean_id = str(token_id).strip()
        if not clean_id:
            return False, "空ID", clean_id

        conn = http.client.HTTPSConnection(API_DOMAIN, timeout=10)
        
        # 构造删除路径 /api/token/{id}/
        # 注意：这里根据您的参考代码逻辑拼接 URL
        path = f"/api/token/{clean_id}/"
        
        conn.request("DELETE", path, "", HEADERS)
        res = conn.getresponse()
        
        # 读取返回内容（可选，用于调试）
        # response_body = res.read().decode("utf-8")

        # 状态码 200 或 204 通常表示删除成功
        if res.status in [200, 201, 204]:
            return True, "删除成功", clean_id
        else:
            return False, f"HTTP {res.status}: {res.reason}", clean_id

    except Exception as e:
        return False, str(e), token_id
    finally:
        if conn:
            conn.close()

def load_ids_from_excel(filepath, column, has_header):
    """从Excel读取ID列表"""
    ids = []
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        # 确定列索引 (A=1, B=2...)
        col_index = ord(column.upper()) - 64
        
        row_count = 0
        for row in ws.iter_rows(min_col=col_index, max_col=col_index, values_only=True):
            row_count += 1
            if has_header and row_count == 1:
                continue
            
            val = row[0]
            if val is not None:
                ids.append(val)
        
        print(f"📄 从文件加载了 {len(ids)} 个待删除对象。")
        return ids
    except Exception as e:
        print(f"❌ 读取Excel失败: {e}")
        return []

# -------------------------- 主程序 --------------------------
if __name__ == "__main__":
    print(f"🚀 开始批量删除任务...")
    
    # 1. 读取 ID
    target_ids = load_ids_from_excel(EXCEL_FILE_PATH, TARGET_COLUMN, HAS_HEADER)
    
    if not target_ids:
        print("没有找到要删除的ID，程序结束。")
        exit()

    success_count = 0
    fail_count = 0
    
    start_time = time.time()

    # 2. 多线程执行删除
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_tasks = {executor.submit(delete_token_task, tid): tid for tid in target_ids}
        
        total = len(target_ids)
        processed = 0

        for future in as_completed(future_tasks):
            processed += 1
            success, msg, tid = future.result()
            
            if success:
                success_count += 1
                # 仅打印简略日志
                print(f"[{processed}/{total}] ✅ 删除成功 ID: {tid}")
            else:
                fail_count += 1
                print(f"[{processed}/{total}] ❌ 删除失败 ID: {tid} | 原因: {msg}")

    # 3. 统计
    duration = time.time() - start_time
    print("=" * 50)
    print(f"任务完成！耗时: {duration:.2f}秒")
    print(f"✅ 成功删除: {success_count}")
    print(f"❌ 失败: {fail_count}")